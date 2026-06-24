from __future__ import annotations

import csv
import json
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_ROOT = Path(r"C:\Users\C2023\Downloads\yuanfudao-knowledge")
DOCS_DIR = SOURCE_ROOT / "documents"
RELEASE_ROOT = Path("release/yuanfudao-course-catalog-cleaned")
VOLCENGINE_ROOT = Path("release/yuanfudao-volcengine-kb-upload")


FIELDS = [
    "record_id",
    "业务线",
    "商品/SKU名称",
    "定价",
    "价格单位",
    "适用人群/年级",
    "课程形式",
    "课时/周期",
    "上课时间/排期",
    "教材/资料/实物",
    "服务/伴学",
    "赠品/权益",
    "后转价格/正价课价格",
    "商品卖点",
    "商品介绍",
    "链接",
    "销转流程",
    "来源文件",
    "来源位置",
    "备注",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\\([.\-+()])", r"\1", text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def clean_inline(value: Any, limit: int | None = None) -> str:
    text = clean_text(value)
    text = re.sub(r"\s*\n\s*", "；", text)
    text = re.sub(r"\s{2,}", " ", text).strip("； ")
    if limit and len(text) > limit:
        return text[: limit - 1] + "…"
    return text


def price_to_text(value: Any) -> str:
    if value is None or str(value).strip() == "":
        return ""
    try:
        num = float(value)
        if num.is_integer():
            return str(int(num))
        return str(num).rstrip("0").rstrip(".")
    except Exception:
        return clean_inline(value)


def add_record(records: list[dict[str, str]], **kwargs: str) -> None:
    record = {field: "" for field in FIELDS}
    record.update({k: clean_text(v) for k, v in kwargs.items() if k in record})
    record["record_id"] = f"YCAT-{len(records) + 1:03d}"
    records.append(record)


def load_source_lines(name: str) -> list[str]:
    path = DOCS_DIR / name
    if not path.exists():
        return []
    return path.read_text(encoding="utf-8", errors="ignore").splitlines()


def source_range(name: str, start: int, end: int) -> str:
    return f"{name}:L{start}-L{end}"


def extract_from_workbook(records: list[dict[str, str]]) -> None:
    path = DOCS_DIR / "猿辅导课程货盘.xlsx"
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    for row_idx in range(2, ws.max_row + 1):
        vals = [ws.cell(row_idx, col).value for col in range(1, 13)]
        if not any(clean_text(v) for v in vals):
            continue
        business, price, _, _, selling, intro, audience, link, idea, value, highlight, flow = vals
        intro_text = clean_text(intro)
        price_match = re.search(r"后转价格[：:\s]*(.+)", intro_text)
        post_price = price_match.group(1).strip() if price_match else ""
        add_record(
            records,
            业务线=business,
            **{
                "商品/SKU名称": clean_inline(business),
                "定价": price_to_text(price),
                "价格单位": "元" if price_to_text(price) else "",
                "适用人群/年级": audience,
                "课程形式": infer_course_format(intro_text + "\n" + clean_text(selling)),
                "课时/周期": infer_lesson_period(intro_text),
                "上课时间/排期": infer_schedule(intro_text),
                "教材/资料/实物": infer_materials(intro_text + "\n" + clean_text(value)),
                "服务/伴学": infer_service(intro_text + "\n" + clean_text(value)),
                "赠品/权益": infer_benefits(intro_text + "\n" + clean_text(value)),
                "后转价格/正价课价格": post_price,
                "商品卖点": selling,
                "商品介绍": intro,
                "链接": link,
                "销转流程": flow,
                "来源文件": "猿辅导课程货盘.xlsx",
                "来源位置": f"Sheet1!R{row_idx}",
                "备注": "来自原始课程货盘表；未出现的字段保持为空。",
            },
        )


def infer_course_format(text: str) -> str:
    parts: list[str] = []
    if "直播" in text:
        parts.append("直播")
    if "录播" in text:
        parts.append("录播")
    if "图书" in text or "书" in text or "教材" in text or "盒子" in text:
        parts.append("实物/图书")
    if "互动录播" in text:
        parts.append("互动录播")
    return " + ".join(dict.fromkeys(parts))


def infer_lesson_period(text: str) -> str:
    patterns = [
        r"\d+\s*天\s*\d+\s*(?:节|次|课时)",
        r"\d+\s*(?:节|次|课时)",
        r"\d+\s*天",
        r"\d+\s*周",
        r"\d+\s*套",
    ]
    hits: list[str] = []
    for pattern in patterns:
        hits.extend(re.findall(pattern, text))
    return "；".join(dict.fromkeys(hits[:6]))


def infer_schedule(text: str) -> str:
    lines = [clean_inline(line) for line in text.splitlines()]
    hits = [
        line
        for line in lines
        if any(key in line for key in ["周", "每天", "19:00", "20:10", "上课", "排期", "早班", "晚班"])
    ]
    return "；".join(dict.fromkeys(hits[:4]))


def infer_materials(text: str) -> str:
    lines = [clean_inline(line) for line in text.splitlines()]
    hits = [
        line
        for line in lines
        if any(key in line for key in ["教材", "资料", "电子", "实物", "图书", "盒子", "书", "礼盒", "押题卷"])
    ]
    return "；".join(dict.fromkeys(hits[:5]))


def infer_service(text: str) -> str:
    lines = [clean_inline(line) for line in text.splitlines()]
    hits = [line for line in lines if any(key in line for key in ["班主任", "伴学", "答疑", "辅导", "老师跟进"])]
    return "；".join(dict.fromkeys(hits[:5]))


def infer_benefits(text: str) -> str:
    lines = [clean_inline(line) for line in text.splitlines()]
    hits = [line for line in lines if any(key in line for key in ["赠", "加赠", "包邮", "福利", "礼"])]
    return "；".join(dict.fromkeys(hits[:5]))


def add_manual_supplements(records: list[dict[str, str]]) -> None:
    add_record(
        records,
        业务线="学科教辅",
        **{
            "商品/SKU名称": "《2026 新课标-全科原型题冲刺》",
            "定价": "16.8",
            "价格单位": "元",
            "适用人群/年级": "在读1-5年级 = 9月升2-6年级，用户下单在读1-5年级，发2-6年级教材",
            "课程形式": "纯图书",
            "课时/周期": "",
            "上课时间/排期": "",
            "教材/资料/实物": "4本书；书籍配套400分钟配套视频讲解 + 70道原型题精讲",
            "服务/伴学": "",
            "赠品/权益": "",
            "后转价格/正价课价格": "",
            "商品卖点": "学原型题，直击学年重难点，一通百通，举一反三；纯图书，无课。",
            "商品介绍": "实付16.8元，4本书，纯图书，无课；扫码即可看配套视频，孩子可在家自学。",
            "来源文件": "【在售】16.8原型题-满分冲刺.md",
            "来源位置": source_range("【在售】16.8原型题-满分冲刺.md", 5, 19),
            "备注": "该来源为2026.05.28调整上新口径；与课程货盘表中的16.8学科教辅直播课口径不同，保留为独立记录。",
        },
    )
    add_record(
        records,
        业务线="学科/研学营",
        **{
            "商品/SKU名称": "杨玉彪老师14天研学营",
            "定价": "19.8",
            "价格单位": "元",
            "适用人群/年级": "在读1-5年级 = 9月升2-6年级，用户下单在读1-5年级，发2-6年级教材",
            "课程形式": "直播 + 资料 + 伴学",
            "课时/周期": "5天10节直播课；1次摸底测；14天伴学",
            "上课时间/排期": "",
            "教材/资料/实物": "8套电子资料；额外加赠16.8元《原型题冲刺》书籍",
            "服务/伴学": "14天三师全程陪跑：主讲 + 辅导老师 + AI助教",
            "赠品/权益": "额外加赠16.8元《原型题冲刺》书籍",
            "后转价格/正价课价格": "",
            "商品卖点": "5天10节直播课 + 1次摸底测 + 8套电子资料 + 14天伴学。",
            "商品介绍": "下单后需看短信，按短信链接填写收货地址并兑换课程，官方来电需接听。",
            "来源文件": "【在售】16.8原型题-满分冲刺.md",
            "来源位置": source_range("【在售】16.8原型题-满分冲刺.md", 15, 23),
            "备注": "来源文档明确为独立商品选项。",
        },
    )
    add_record(
        records,
        业务线="新奥数/思维",
        **{
            "商品/SKU名称": "29元《秒懂新奥数》",
            "定价": "29",
            "价格单位": "元",
            "适用人群/年级": "招生年级：在读1-5年级 = 9月升2-6年级；目标人群含1-6年级、校内80分以上想提升数学思维/应对新考法/考试拿高分学生，及4-6年级想拓展奥数/小升初冲刺学生",
            "课程形式": "实物图书 + 直播课 + 电子资料 + 伴学",
            "课时/周期": "5天10节名师课；14天班主任伴学",
            "上课时间/排期": "5天课程：周一到周五每天晚上19:00；5天10课时直播课程，每节课30分钟，一天2节",
            "教材/资料/实物": "6本书：32种新奥解题策略、计算/应用/几何、语数英期中期末押题卷、新教材新考法早知道；加赠8套语数英全科电子资料",
            "服务/伴学": "班主任/老师服务；下单48h内班主任联系；14天班主任全程答疑伴学",
            "赠品/权益": "完课加赠实物礼，随机送1件实物礼品，全国包邮到家；斑马百科体验会员",
            "后转价格/正价课价格": "",
            "商品卖点": "数学专项浅奥方向 + 校内语文英语新考法及押题卷；购买盒子额外加赠5天10节特训课。",
            "商品介绍": "6本图书约320个视频、约700分钟；搭配5天10节名师直播课和班主任服务。",
            "来源文件": "【在售】29元《秒懂新奥数》介绍.md",
            "来源位置": source_range("【在售】29元《秒懂新奥数》介绍.md", 1, 549),
            "备注": "价格来自文件名；课程/赠品/人群来自正文明确描述。",
        },
    )
    add_record(
        records,
        业务线="新奥数/思维",
        **{
            "商品/SKU名称": "29元新奥数21天学习内容",
            "定价": "29",
            "价格单位": "元",
            "适用人群/年级": "",
            "课程形式": "直播 + 带练 + 答疑 + 打卡 + 实物",
            "课时/周期": "21天；6次（3次核心直播课、3次主讲带练课）12课时清北名师领衔精讲直播课",
            "上课时间/排期": "",
            "教材/资料/实物": "24种奥数思维模型；降维覆盖96个校内重难点；100类经典奥数题型带练；1次高水平思维测评 + 1次全国联考",
            "服务/伴学": "21天金牌班主任全程伴学；辅导老师定期1v1、1v3答疑课",
            "赠品/权益": "",
            "后转价格/正价课价格": "",
            "商品卖点": "围绕计算、图形、思维、应用四个模块，课前预习、课堂学习、课后练习闭环。",
            "商品介绍": "21天学习内容包含核心课、带练课、答疑课、周打卡、摸底测/导学课和全国联考。",
            "来源文件": "29元新奥数产品介绍.md",
            "来源位置": source_range("29元新奥数产品介绍.md", 1, 49),
            "备注": "可能与货盘表中29元思维产品存在业务重叠；因来源口径不同，保留独立记录。",
        },
    )
    add_record(
        records,
        业务线="学科教辅",
        **{
            "商品/SKU名称": "59元品《全科满分营-原型题秒解重难考点》",
            "定价": "59",
            "价格单位": "元",
            "适用人群/年级": "当前处于升年级阶段，招生年级：在读1-5年级 = 9月升2-6年级，发2-6年级教材；目标用户：有支付力、有提分需求、希望寻求好方法提效",
            "课程形式": "实物盒子 + 直播课 + 录播/视频 + 电子资料 + 伴学",
            "课时/周期": "10节核心清北名师直播课；198节语数英全科核心知识精讲课/视频，合计约2290分钟；5天10课时直播课程",
            "上课时间/排期": "5天课程：周一到周五每天晚上19:00；每节课30分钟，一天2节",
            "教材/资料/实物": "26本 + 16套押题卷 + 2张挂图 + 1套桌面台历；加赠8套语数英全科电子资料",
            "服务/伴学": "班主任/老师服务；14天班主任全程答疑伴学",
            "赠品/权益": "购买盒子额外加赠5天10节特训课 + 198节语数英全科核心知识精讲课",
            "后转价格/正价课价格": "",
            "商品卖点": "原型题秒解重难考点；用主书同步学校教材单元学习，押题卷+原型题大招辞典备考。",
            "商品介绍": "全科满分营包含语数英专项书籍、押题卷、挂图、台历、直播课和录播视频资源。",
            "来源文件": "【在售】59元品《全科满分营-原型题秒解重难考点》.md",
            "来源位置": source_range("【在售】59元品《全科满分营-原型题秒解重难考点》.md", 1, 507),
            "备注": "价格来自文件名；未见链接字段。",
        },
    )
    add_record(
        records,
        业务线="学科纯课",
        **{
            "商品/SKU名称": "5元达播纯课",
            "定价": "5",
            "价格单位": "元",
            "适用人群/年级": "2-6年级可拍",
            "课程形式": "直播 + 录播赠课 + 电子资料 + 完课礼",
            "课时/周期": "产品：11课时清北名师线上直播课；PS：13课时=5天9课时数学+1天2课时语文+2课时英语录播；额外加赠6节思维录播课",
            "上课时间/排期": "每次19:00开始上课；英语体验课赠课day3晚20:30解锁；加赠6节思维录播课每周一下午",
            "教材/资料/实物": "电子资料包含试卷和知识点汇总；资料均为电子版，报名后添加班主任微信领取",
            "服务/伴学": "学情规划诊断 + 金牌辅导老师专项服务",
            "赠品/权益": "限时加赠小猿手提编织袋；前4节直播出勤时长均不低于80%赠送实物到课礼，结课后7个工作日内包邮到家，礼品随机",
            "后转价格/正价课价格": "",
            "商品卖点": "语数英核心课程，学情规划诊断，资料和完课礼结合。",
            "商品介绍": "直播售纯课需介绍课程节数、领课和资料；课程支持回放，3年内无限次观看。",
            "来源文件": "5 元达播纯课.md",
            "来源位置": source_range("5 元达播纯课.md", 11, 143),
            "备注": "来源内同时出现11课时和13课时口径，按原文保留。",
        },
    )
    add_record(
        records,
        业务线="奥数思维",
        **{
            "商品/SKU名称": "49元奥数盒子",
            "定价": "49",
            "价格单位": "元",
            "适用人群/年级": "",
            "课程形式": "直播 + 实物教材 + 伴学",
            "课时/周期": "12课时名师精讲直播 + 5本校内校外名题精练",
            "上课时间/排期": "",
            "教材/资料/实物": "5本校内校外名题精练",
            "服务/伴学": "金牌班主任全程伴学",
            "赠品/权益": "",
            "后转价格/正价课价格": "",
            "商品卖点": "奥数入门科学规划，串联计算应用核心模块20+难点考点。",
            "商品介绍": "49元包含12课时名师精讲直播、5本校内校外名题精练、金牌班主任全程伴学。",
            "来源文件": "49元奥数盒子.md",
            "来源位置": source_range("49元奥数盒子.md", 1, 5),
            "备注": "未在来源中找到明确适用年级，留空。",
        },
    )
    add_record(
        records,
        业务线="奥数思维",
        **{
            "商品/SKU名称": "269思维盒子",
            "定价": "269",
            "价格单位": "元",
            "适用人群/年级": "",
            "课程形式": "直播 + 实物教材 + 伴学",
            "课时/周期": "16课时名师精讲直播 + 20册新奥数名题精练 + 30天全程伴学；8次16课时清北名师领衔精讲直播课",
            "上课时间/排期": "晚班：周五&周六晚19:00；早班：周六&周日早10:00",
            "教材/资料/实物": "20册新奥数名题精练",
            "服务/伴学": "30天金牌班主任全程伴学",
            "赠品/权益": "",
            "后转价格/正价课价格": "",
            "商品卖点": "新奥数名题精练结合名师直播和全程伴学。",
            "商品介绍": "269元=16课时名师精讲直播 + 20册新奥数名题精练 + 30天全程伴学。",
            "来源文件": "269思维盒子.md",
            "来源位置": source_range("269思维盒子.md", 31, 127),
            "备注": "未在来源中找到明确适用年级，留空。",
        },
    )
    add_record(
        records,
        业务线="学科/语数双科",
        **{
            "商品/SKU名称": "语数双科纯课特训营",
            "定价": "",
            "价格单位": "",
            "适用人群/年级": "",
            "课程形式": "直播 + 英语录播赠课 + 思维录播赠课 + 电子资料 + 完课礼",
            "课时/周期": "13课时=5天9课时数学+1天2课时语文+2课时英语录播；额外加赠6节思维录播课",
            "上课时间/排期": "每次19:00开始上课；英语体验课赠课day3晚20:30解锁；额外加赠6节思维录播课每周一下午",
            "教材/资料/实物": "下学期电子资料；资料均为电子版，报名后添加班主任微信发放",
            "服务/伴学": "班主任答疑伴学；课中老师监督上课；课后整理课程报告",
            "赠品/权益": "前4节直播出勤时长均不低于80%赠送实物到课礼；结课后7个工作日内包邮到家，礼品随机",
            "后转价格/正价课价格": "",
            "商品卖点": "学练考三位一体解决校内问题，支持3年内无限次回放。",
            "商品介绍": "语数双科纯课特训营包含数学、语文直播及英语/思维赠课，配套电子资料和完课礼规则。",
            "来源文件": "语数双科纯课特训营介绍.md",
            "来源位置": source_range("语数双科纯课特训营介绍.md", 5, 144),
            "备注": "来源未明确价格和适用年级，留空。",
        },
    )
    add_record(
        records,
        业务线="自然拼读正价班",
        **{
            "商品/SKU名称": "自然拼读正价班（一、二年级0基础）",
            "定价": "2380",
            "价格单位": "元",
            "适用人群/年级": "一、二年级0基础",
            "课程形式": "正价班课",
            "课时/周期": "5节寒假课程 + 16节春季课程；每节正式课包含两个课时，30min-10min-30min",
            "上课时间/排期": "",
            "教材/资料/实物": "",
            "服务/伴学": "",
            "赠品/权益": "",
            "后转价格/正价课价格": "2380",
            "商品卖点": "课程收获：26字母能读能写；CVC单词拼读；50+核心词汇；听口启蒙；阅读量600+。",
            "商品介绍": "自然拼读正价班低年级0基础口径。",
            "来源文件": "自然拼读-产品培训文档-26.3.11.md",
            "来源位置": source_range("自然拼读-产品培训文档-26.3.11.md", 109, 117),
            "备注": "正价班课价格，非低价体验课。",
        },
    )
    for sku, audience, gain in [
        ("自然拼读正价班（一二年级有基础）", "一二年级有基础", "70%小学自拼，累计词汇量300，28组校内+剑桥句型，阅读量1500+"),
        ("自然拼读正价班（三年级）", "三年级", "70%小学自拼，3下校内主题词汇、句型，3下听说读写题型训练"),
        ("自然拼读正价班（四年级）", "四年级", "自拼综合应用，4下校内主题词汇、句型，6大语法专项，4下听说读写技能训练"),
    ]:
        add_record(
            records,
            业务线="自然拼读正价班",
            **{
                "商品/SKU名称": sku,
                "定价": "2880",
                "价格单位": "元",
                "适用人群/年级": audience,
                "课程形式": "正价班课",
                "课时/周期": "7节寒假课程 + 16节春季课程；每节正式课包含2.5课时，25min核心课-10min休息-25min核心课-10min休息-15min核心课",
                "上课时间/排期": "",
                "教材/资料/实物": "",
                "服务/伴学": "",
                "赠品/权益": "",
                "后转价格/正价课价格": "2880",
                "商品卖点": f"课程收获：{gain}",
                "商品介绍": "自然拼读正价班有基础/高年级口径。",
                "来源文件": "自然拼读-产品培训文档-26.3.11.md",
                "来源位置": source_range("自然拼读-产品培训文档-26.3.11.md", 119, 133),
                "备注": "正价班课价格，非低价体验课。",
            },
        )


def write_csv(records: list[dict[str, str]], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(records)


def write_jsonl(records: list[dict[str, str]], path: Path) -> None:
    with path.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def md_escape(value: str) -> str:
    return clean_inline(value).replace("|", "\\|")


def write_markdown(records: list[dict[str, str]], path: Path) -> None:
    lines: list[str] = []
    lines.append("# 猿辅导课程货盘清洗版")
    lines.append("")
    lines.append(f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- 来源目录：`{SOURCE_ROOT}`")
    lines.append("- 清洗原则：只保留源文件明确写出的字段；未找到就留空；不同来源口径不强行合并。")
    lines.append("- 使用提醒：价格、排期、权益、赠品、活动有效期等强时效信息，入库后仍建议以最新活动页、班主任通知或后台为准。")
    lines.append("")
    lines.append("## 总表")
    lines.append("")
    table_fields = ["record_id", "业务线", "商品/SKU名称", "定价", "适用人群/年级", "课程形式", "课时/周期", "来源文件", "备注"]
    lines.append("|" + "|".join(table_fields) + "|")
    lines.append("|" + "|".join(["---"] * len(table_fields)) + "|")
    for record in records:
        lines.append("|" + "|".join(md_escape(record[field]) for field in table_fields) + "|")
    lines.append("")
    lines.append("## 明细")
    for record in records:
        lines.append("")
        lines.append(f"### {record['record_id']} {record['商品/SKU名称'] or record['业务线']}")
        for field in FIELDS:
            if field == "record_id":
                continue
            value = clean_text(record[field])
            if value:
                lines.append(f"- **{field}**：{value}")
            else:
                lines.append(f"- **{field}**：")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_txt(records: list[dict[str, str]], path: Path) -> None:
    chunks: list[str] = []
    for record in records:
        title = f"[{record['record_id']}] {record['商品/SKU名称'] or record['业务线']}"
        chunks.append(title)
        chunks.append("=" * len(title))
        for field in FIELDS:
            if field == "record_id":
                continue
            value = clean_inline(record[field])
            chunks.append(f"{field}: {value}" if value else f"{field}:")
        chunks.append("")
    path.write_text("\n".join(chunks), encoding="utf-8")


def write_readme(path: Path, count: int) -> None:
    path.write_text(
        f"""# 猿辅导课程货盘清洗版

本目录由 `C:\\Users\\C2023\\Downloads\\yuanfudao-knowledge` 中的课程货盘表和明确产品资料清洗生成。

## 文件说明

- `yuanfudao_course_catalog_ready.md`：推荐上传到知识库的 Markdown 版本。
- `yuanfudao_course_catalog_for_kb.txt`：纯文本导入版。
- `yuanfudao_course_catalog_entries.csv`：结构化表格版。
- `yuanfudao_course_catalog_entries.jsonl`：结构化 JSONL 版。

## 清洗原则

- 共输出 {count} 条课程/商品记录。
- 只抽取源文件明确写出的事实字段。
- 未找到的字段留空。
- 不同来源存在不同口径时，不强行合并，保留为独立记录并在备注中说明。

## 当前风险提示

价格、排期、权益、赠品、活动有效期等信息具有强时效性，上传知识库后仍建议以最新活动页、班主任通知或后台为准。
""",
        encoding="utf-8",
    )


def zip_dir(root: Path, zip_path: Path) -> None:
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file in root.rglob("*"):
            if file.is_file():
                zf.write(file, file.relative_to(root.parent).as_posix())


def sync_to_volcengine_package() -> None:
    if not VOLCENGINE_ROOT.exists():
        return
    course_dir = VOLCENGINE_ROOT / "05_course_catalog"
    if course_dir.exists():
        shutil.rmtree(course_dir)
    course_dir.mkdir(parents=True, exist_ok=True)
    for name in [
        "README_课程货盘说明.md",
        "yuanfudao_course_catalog_ready.md",
        "yuanfudao_course_catalog_for_kb.txt",
        "yuanfudao_course_catalog_entries.csv",
        "yuanfudao_course_catalog_entries.jsonl",
    ]:
        src = RELEASE_ROOT / name
        if src.exists():
            shutil.copy2(src, course_dir / name)
    upload_md = VOLCENGINE_ROOT / "01_markdown_upload" / "04_猿辅导课程货盘知识库.md"
    upload_md.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(RELEASE_ROOT / "yuanfudao_course_catalog_ready.md", upload_md)
    upload_txt = VOLCENGINE_ROOT / "02_txt_upload" / "03_猿辅导课程货盘导入版.txt"
    upload_txt.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(RELEASE_ROOT / "yuanfudao_course_catalog_for_kb.txt", upload_txt)
    (VOLCENGINE_ROOT / "README_火山知识库上传说明.md").write_text(
        """# 猿辅导清洗知识库 - 平台无关上传包

这个压缩包只包含可放入火山引擎或其他知识库的清洗数据，不包含项目模板、manifest、向量库、本地数据库或项目配置。

## 推荐上传顺序

1. `01_markdown_upload/00_猿辅导FAQ_话术_SOP总知识库.md`
   - FAQ、销售话术、群发 SOP 总文档。
2. `01_markdown_upload/04_猿辅导课程货盘知识库.md`
   - 课程货盘/产品事实清洗版。
3. 或者分别上传：
   - `01_markdown_upload/01_猿辅导FAQ问答知识库.md`
   - `01_markdown_upload/02_猿辅导销售话术知识库.md`
   - `01_markdown_upload/03_猿辅导1天2次群发SOP知识库.md`
   - `01_markdown_upload/04_猿辅导课程货盘知识库.md`
4. 如果平台对 Markdown 解析不好，可以上传 `02_txt_upload/` 里的 TXT 文件。
5. `03_by_product_markdown/` 是按产品拆分的 FAQ/话术/SOP Markdown。
6. `04_structured_data/` 是 FAQ/话术/SOP 的 CSV/JSONL 结构化版本。
7. `05_course_catalog/` 是课程货盘的 Markdown、TXT、CSV、JSONL 版本。

## 当前覆盖范围

- FAQ 与销售话术：311 条清洗记录
- 1天2次群发 SOP：98 条清洗记录
- 课程货盘/产品事实：21 条课程或商品记录
- 内容范围：自然拼读、剑桥英语、学科、英语、数学/思维、通用问题、群发 SOP、课程货盘

## 当前不包含

- 图片/视频素材本身
- 本地向量索引或数据库
- 未在源文件中明确出现的字段

涉及实时价格、权益、售后、排期时，建议以最新活动页、班主任通知或后台为准。
""",
        encoding="utf-8",
    )
    zip_dir(VOLCENGINE_ROOT, Path("release/yuanfudao-volcengine-kb-upload.zip"))


def main() -> None:
    records: list[dict[str, str]] = []
    extract_from_workbook(records)
    add_manual_supplements(records)

    if RELEASE_ROOT.exists():
        shutil.rmtree(RELEASE_ROOT)
    RELEASE_ROOT.mkdir(parents=True, exist_ok=True)

    write_markdown(records, RELEASE_ROOT / "yuanfudao_course_catalog_ready.md")
    write_txt(records, RELEASE_ROOT / "yuanfudao_course_catalog_for_kb.txt")
    write_csv(records, RELEASE_ROOT / "yuanfudao_course_catalog_entries.csv")
    write_jsonl(records, RELEASE_ROOT / "yuanfudao_course_catalog_entries.jsonl")
    write_readme(RELEASE_ROOT / "README_课程货盘说明.md", len(records))
    zip_dir(RELEASE_ROOT, Path("release/yuanfudao-course-catalog-cleaned.zip"))
    sync_to_volcengine_package()

    print(
        json.dumps(
            {
                "records": len(records),
                "output_dir": str(RELEASE_ROOT),
                "zip": "release/yuanfudao-course-catalog-cleaned.zip",
                "volcengine_zip": "release/yuanfudao-volcengine-kb-upload.zip"
                if VOLCENGINE_ROOT.exists()
                else "",
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
